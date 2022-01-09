from flask import Flask
from flask import request
from openpyxl import load_workbook
import mygene
import re
from datetime import date
import glob, os
from os.path import exists


_upload_dir = "./files/"
ref_folder = "./reference/"
output_folder = "./homedir/"
app = Flask(__name__)


def get_score_genes(patient_id, score_map, score_suffix, nrzThreshold, crcThreshold, lowCRCThreshold):
    data_out = {}
    mg = mygene.MyGeneInfo()
    ens = []
    ens_nrz = {}
    ens_crc = {}

    f = open(_upload_dir + patient_id + score_suffix)
    f.readline()
    for line in f:
        line_vec = line.split('\t')
        if len(line_vec)>2:
            if abs(float(line_vec[1])) > nrzThreshold or float(line_vec[2]) > crcThreshold:
                ens_nrz[line_vec[0]] = float(line_vec[1])
                ens_crc[line_vec[0]] = float(line_vec[2])
                ens.append(line_vec[0])
                if line_vec[0] == "ENSG00000099942":
                    print(line);

    mg = mygene.MyGeneInfo()
    g_info = mg.querymany(ens, scopes='ensembl.gene')
    for gene in g_info:
        if "symbol" not in gene:
            continue
        if gene["symbol"].upper() not in score_map:
            continue
       # print(gene['symbol'] + "\t" + str(ens_nrz[gene["query"]]) + "\t" + str(ens_crc[gene["query"]]))
        z_score = ens_nrz[gene["query"]]
        crc_score = ens_crc[gene["query"]]
        rules = score_map[gene['symbol']]
        for ind_rule in rules:
            type = ind_rule["aberration_value"]
            if type == "over":
                if z_score > nrzThreshold:
                    print("nrz: " + gene['symbol'] + "\t" + str(ind_rule["drug"]))
                if crc_score > crcThreshold:
                    print("crc: " + gene['symbol'] + "\t" + str(ind_rule["drug"]))
                if (z_score > nrzThreshold or crc_score > crcThreshold):
                    if ind_rule["drug"] not in data_out:
                        data_out[ind_rule["drug"]] = []
                    table_entry = {}
                    table_entry["gene"] = gene["symbol"]
                    table_entry["z_score"] = z_score
                    table_entry["crc"] = crc_score
                    table_entry["sense"] = ind_rule["sense"]
                    table_entry["direction"] = "over"
                    data_out[ind_rule["drug"]].append(table_entry)
            elif type == "under":
                if (-1*z_score > nrzThreshold or crc_score < lowCRCThreshold):
                    if ind_rule["drug"] not in data_out:
                        data_out[ind_rule["drug"]] = []
                    table_entry = {}
                    table_entry["gene"] = gene["symbol"]
                    table_entry["z_score"] = z_score
                    table_entry["crc"] = crc_score
                    table_entry["sense"] = ind_rule["sense"]
                    table_entry["direction"] = "under"
                    data_out[ind_rule["drug"]].append(table_entry)
    return data_out

def get_snp_genes(patient_id, snv_map,snv_suffix,):
    d = {'CYS': 'C', 'ASP': 'D', 'SER': 'S', 'GLN': 'Q', 'LYS': 'K',
         'ILE': 'I', 'PRO': 'P', 'THR': 'T', 'PHE': 'F', 'ASN': 'N',
         'GLY': 'G', 'HIS': 'H', 'LEU': 'L', 'ARG': 'R', 'TRP': 'W',
         'ALA': 'A', 'VAL': 'V', 'GLU': 'E', 'TYR': 'Y', 'MET': 'M', 'XAA': 'X'}

    ens_snv = {}
    data_out = {}
    f = open(_upload_dir + patient_id + snv_suffix)
    f.readline()
    proc = False
    ens = []
    for line in f:
        if "#CHROM\tPOS" not in line and not proc:
            continue;
        elif not proc:
            proc = True
            continue
        val = re.search(r"ANN.*?\|\|", line)
        match_val = ""
        if val is not None:
            match_val = val.group(0)
        if match_val != "":
            match_vec = match_val.split('|')
#            print(match_vec)
            variant_type = ""
            if len(match_vec) > 1:
                variant_type = match_vec[1].lower().strip()
                if variant_type == "missense_variant":
                    ens_gene_re = re.search(r"ENSG.\d*", match_val)
                   # print(ens_gene_re)
                    if ens_gene_re is not None:
                        ens_gene = ens_gene_re.group(0)
                        ens.append(ens_gene)
                  #      print(match_val)
                        aa_change_re = re.search("p\..*?\|", match_val)
                  #      print(aa_change_re)
                        aa_change = ""
                        if aa_change_re is not None:
                            aa_change = aa_change_re.group(0).replace("|", "").replace("p.","").upper()
                            if ens_gene not in ens_snv:
                                ens_snv[ens_gene] = []
                            for abbrev in d:
                                aa_change = aa_change.replace(abbrev.upper(),d[abbrev].upper())
                            ens_snv[ens_gene].append(aa_change)

    mg = mygene.MyGeneInfo()
    g_info = mg.querymany(ens, scopes='ensembl.gene')
    for gene in g_info:
        if "symbol" not in gene:
            continue
        if gene["symbol"].upper() not in snv_map:
            continue
            # print(gene['symbol'] + "\t" + str(ens_nrz[gene["query"]]) + "\t" + str(ens_crc[gene["query"]]))
        for aa_change_loc in ens_snv[gene["query"]]:
            rules = snv_map[gene['symbol']]
            for ind_rule in rules:
                if ind_rule["drug"] not in data_out:
                    data_out[ind_rule["drug"]] = []
                table_entry = {}
                table_entry["gene"] = gene["symbol"]
                table_entry["aa_change"] = aa_change_loc
                table_entry["sense"] = ind_rule["sense"]
                data_out[ind_rule["drug"]].append(table_entry)
    return data_out

def get_snp_genes_peach(patient_id, snv_map):
    d = {'CYS': 'C', 'ASP': 'D', 'SER': 'S', 'GLN': 'Q', 'LYS': 'K',
         'ILE': 'I', 'PRO': 'P', 'THR': 'T', 'PHE': 'F', 'ASN': 'N',
         'GLY': 'G', 'HIS': 'H', 'LEU': 'L', 'ARG': 'R', 'TRP': 'W',
         'ALA': 'A', 'VAL': 'V', 'GLU': 'E', 'TYR': 'Y', 'MET': 'M', 'XAA': 'X'}

    ens_snv = {}
    data_out = {}
    xlsx_file = _upload_dir + patient_id + "_beatCC_Somatic.xlsx"
    if not exists(xlsx_file):
        return data_out;
    workbook = load_workbook(xlsx_file)
    worksheet = workbook.active
    row_counter = 0
    for row_cells in worksheet.iter_rows():
        if row_counter == 0:
            row_counter = row_counter + 1
            continue
        variant_type = str(row_cells[4].value).upper().strip()
        if "MISSENSE" in variant_type:
            aa_change = str(row_cells[6].value)
            ens_gene = str(row_cells[5].value)
            if ens_gene not in ens_snv:
                ens_snv[ens_gene] = []
            ens_snv[ens_gene].append(aa_change)
    print("NRAS")
    print(snv_map["NRAS"])
    print(ens_snv)
    for gene in ens_snv:
        if gene.upper() not in snv_map:
            continue
        for aa_change in ens_snv[gene]:
            rules = snv_map[gene]
            for ind_rule in rules:
                if ind_rule["drug"] not in data_out:
                    data_out[ind_rule["drug"]] = []
                table_entry = {}
                table_entry["gene"] = gene
                table_entry["aa_change"] = aa_change
                table_entry["sense"] = ind_rule["sense"]
                data_out[ind_rule["drug"]].append(table_entry)
    return data_out

def get_cnv_genes(patient_id, cnv_map, cnv_suffix,cnv_threshold):
    f = open(_upload_dir + patient_id + cnv_suffix)
    f.readline()
    proc = False
    ens = []
    data_out = {}
    for line in f:
        if "#CHROM\tPOS" not in line and not proc:
            continue;
        elif not proc:
            proc = True
            continue
        line_vec = line.split("\t")
        gene = []
        if len(line_vec) > 7:
            int_line_vec = line_vec[7].split(";")
            log_fold_change = float(int_line_vec[4].replace("LOG2FC=","").strip())
            gene = int_line_vec[5].replace("GENE=","")
            print(gene + "\t" + str(log_fold_change))
            if gene not in cnv_map:
                continue
            if log_fold_change < cnv_threshold:
                continue
            rules = cnv_map[gene]
            for ind_rule in rules:
                if ind_rule["drug"] not in data_out:
                    data_out[ind_rule["drug"]] = []
                table_entry = {}
                table_entry["gene"] = gene
                table_entry["log_fold_change"] = log_fold_change
                table_entry["sense"] = ind_rule["sense"]
                data_out[ind_rule["drug"]].append(table_entry)
    return(data_out)

def get_fusion_genes(patient_id, fusion_map, fusion_suffix):
    f = open(_upload_dir + patient_id + fusion_suffix)
    blah = f.read()
    data_out = {}
    blah = blah.replace("\n","**")
    all_matches = re.findall(r"fusion_name=.*?;",blah)
    for fusion in all_matches:
        out_val = fusion.replace("--","_").replace("- -","_").replace("fusion_name=","").replace(";","").strip()
        print (out_val)
        gene_vec = out_val.split("_")
        if len(gene_vec) < 2:
            continue
        gene1 = gene_vec[0]
        gene2 = gene_vec[1]
        print (gene1 + "\t" +  gene2)
        if gene1 not in fusion_map or gene2 not in fusion_map:
            continue
        rules = fusion_map[gene1]
        for ind_rule in rules:
            if ind_rule["gene2"] != gene2:
                continue
            if ind_rule["drug"] not in data_out:
                data_out[ind_rule["drug"]] = []
            table_entry = {}
            table_entry["gene"] = gene1
            table_entry["fusion"] = out_val
            table_entry["sense"] = ind_rule["sense"]
            data_out[ind_rule["drug"]].append(table_entry)
        rules = fusion_map[gene2]
        for ind_rule in rules:
            if ind_rule["gene2"] != gene1:
                continue
            if ind_rule["drug"] not in data_out:
                data_out[ind_rule["drug"]] = []
            table_entry = {}
            table_entry["gene"] = gene1
            table_entry["fusion"] = out_val
            table_entry["sense"] = ind_rule["sense"]
            data_out[ind_rule["drug"]].append(table_entry)
    return(data_out)

def write_html_file(patient):
    filename = output_folder + patient +".html"
    f = open (filename,"w")
    f.write("<!DOCTYPE html>\n")
    f.write("<html>\n")
    f.write("\t<body>\n")
    f.write("<div class = \"header\">\n")
    f.write("<img src = \"/static/image001.png\" style=\"display: inline;\"></img><br><br>")
    f.write("<h2 style=\"display: inline; vertical-align:middle;  padding-left: 60px;\">Molecular Guided Report</h2></pre>\n")
    f.write("</div>\n")
    f.write("\t\t<table width=\"800\" border=\"0\">\n")
    f.write("\t\t<tr>\n")
    f.write("\t\t<th style=\"font-family:Georgia, Garamond, Serif;color:white;\"  border=\"0\" bgcolor=\"#4267B2\" align=left>Patient Info</th>\n")
    f.write("\t\t </tr>\n")
    f.write("\t\t</table>\n")
    f.write("<pre><h4> Study ID:       " + patient  + "</h3></pre>\n")
    today = date.today()
    print("Today's date:", today)
    f.write("<pre><h4> Report Date:    " + str(today) + "</h3></pre>\n")
    f.write("<pre><h4> Report Version: 1.0</h3></pre>\n")
    f.write("\t\t</br>\n")
    f.write("\t\t</br>\n")
    f.write("\t\t</br>\n")
    f.write("\t\t<table width=\"800\" border=\"0\">\n")
    f.write("\t\t<tr>\n")
    f.write("\t\t<th style=\"font-family:Georgia, Garamond, Serif;color:white;\"  border=\"0\" bgcolor=\"#4267B2\" align=left>Drug Sensitivity Report</th>\n")
    f.write("\t\t </tr>\n")
    f.write("\t\t</table>\n")
    f.flush()
    f.close()
    return filename

def process_rules_plan(workbook):
    exp_gene = {}
    cnv_gene = {}
    fusion_gene = {}
    snv_gene = {}
    worksheet = workbook['Main']
    output = {}
    for row_cells in worksheet.iter_rows():
        print(row_cells[0].value)
        if row_cells[0].value is None or str(row_cells[0].value).upper() == "DRUG":
            continue
        rule = {}
        rule["drug"] = str(row_cells[0].value).strip()
        gene1 = ""
        gene2 = ""

        gene = str(row_cells[2].value).upper()
        gene1 = gene
        if "_" in gene:
            gene_vec = gene.split("_")
            gene1 = str(gene_vec[0])
            gene2 = str(gene_vec[1])
        rule["gene1"] = gene1
        rule["gene2"] = gene2
        rule["sense"] = str(row_cells[5].value).upper()
        rule["aberration_value"] = str(row_cells[4].value).lower().strip()
        if row_cells[6].value is not None:
            rule["refs"] = "https://www.ncbi.nlm.nih.gov/pubmed/report=" + str(row_cells[6].value).replace(" ", "") \
                           + "medline&format=text"
        else:
            rule["refs"] = ""
        type_v = str(row_cells[3].value).lower().strip()

        if type_v == "none":
            if gene2 == "":
                print("Error in fusion rule: " + gene1)
                continue
            if gene1 not in fusion_gene:
                fusion_gene[gene1] = []
            if gene2 not in fusion_gene:
                fusion_gene[gene2] = []
            fusion_gene[gene1].append(rule)
            fusion_gene[gene2].append(rule)
        elif type_v == "exp":
            if gene1 not in exp_gene:
                exp_gene[gene1] = []
            exp_gene[gene1].append(rule)
        elif type_v == "cnv":
            if gene1 not in cnv_gene:
                cnv_gene[gene1] = []
            cnv_gene[gene1].append(rule)
        elif type_v == "snv":
            if gene1 not in snv_gene:
                snv_gene[gene1] = []
            snv_gene[gene1].append(rule)
    output["snv_gene"] = snv_gene
    output["cnv_gene"] = cnv_gene
    output["fusion_gene"] = fusion_gene
    output["exp_gene"] = exp_gene
    return output


def process_rules_mgt9(workbook):
    exp_gene = {}
    cnv_gene = {}
    fusion_gene = {}
    snv_gene = {}
    worksheet = workbook['Main']
    output = {}
    for row_cells in worksheet.iter_rows():
        type = ""
        print(row_cells[0].value)
        if row_cells[0].value is None or str(row_cells[0].value).upper() == "DRUG":
            continue
        rule = {}
        rule["drug"] = str(row_cells[0].value).strip()
        gene1 = ""
        gene2 = ""

        gene = str(row_cells[2].value).upper()
        gene1 = gene
        if "_" in gene:
            gene_vec = gene.split("_")
            gene1 = str(gene_vec[0])
            gene2 = str(gene_vec[1])

        rule["gene1"] = gene1
        rule["gene2"] = gene2
        rule["sense"] = str(row_cells[5].value).upper()
        rule["aberration_value"] = str(row_cells[4].value).lower().strip()
        if row_cells[6].value is not None:
            rule["refs"] = "https://www.ncbi.nlm.nih.gov/pubmed/report=" + str(row_cells[6].value).replace(" ", "") \
                           + "medline&format=text"
        else:
            rule["refs"] = ""
        type_v = str(row_cells[3].value).lower().strip()

        if type_v == "none":
            if gene2 == "":
                print("Error in fusion rule: " + gene1)
                continue
            if gene1 not in fusion_gene:
                fusion_gene[gene1] = []
            if gene2 not in fusion_gene:
                fusion_gene[gene2] = []
            fusion_gene[gene1].append(rule)
            fusion_gene[gene2].append(rule)
        elif type_v == "exp":
            if gene1 not in exp_gene:
                exp_gene[gene1] = []
            exp_gene[gene1].append(rule)
        elif type_v == "cnv":
            if gene1 not in cnv_gene:
                cnv_gene[gene1] = []
            cnv_gene[gene1].append(rule)
        elif type_v == "snv":
            if gene1 not in snv_gene:
                snv_gene[gene1] = []
            snv_gene[gene1].append(rule)
    output["snv_gene"] = snv_gene
    output["cnv_gene"] = cnv_gene
    output["fusion_gene"] = fusion_gene
    output["exp_gene"] = exp_gene
    return output


def process_rules_peach(workbook):
    exp_gene = {}
    cnv_gene = {}
    fusion_gene = {}
    snv_gene = {}
    worksheet = workbook['Main']
    output = {}
    for row_cells in worksheet.iter_rows():
        type = ""
        print(row_cells[0].value)
        if row_cells[0].value is None or str(row_cells[0].value).upper() == "DRUG":
            continue
        rule = {}
        rule["drug"] = str(row_cells[0].value).strip()
        gene1 = ""
        gene2 = ""

        gene = str(row_cells[2].value).upper()
        gene1 = gene
        if "_" in gene:
            gene_vec = gene.split("_")
            gene1 = str(gene_vec[0])
            gene2 = str(gene_vec[1])

        rule["gene1"] = gene1
        rule["gene2"] = gene2
        rule["sense"] = str(row_cells[5].value).upper()
        rule["aberration_value"] = str(row_cells[4].value).lower().strip()
        if row_cells[6].value is not None:
            rule["refs"] = "https://www.ncbi.nlm.nih.gov/pubmed/report=" + str(row_cells[6].value).replace(" ", "") \
                           + "medline&format=text"
        else:
            rule["refs"] = ""
        type_v = str(row_cells[3].value).lower().strip()

        if type_v == "none":
            if gene2 == "":
                print("Error in fusion rule: " + gene1)
                continue
            if gene1 not in fusion_gene:
                fusion_gene[gene1] = []
            if gene2 not in fusion_gene:
                fusion_gene[gene2] = []
            fusion_gene[gene1].append(rule)
            fusion_gene[gene2].append(rule)
        elif type_v == "exp":
            if gene1 not in exp_gene:
                exp_gene[gene1] = []
            exp_gene[gene1].append(rule)
        elif type_v == "cnv":
            if gene1 not in cnv_gene:
                cnv_gene[gene1] = []
            cnv_gene[gene1].append(rule)
        elif type_v == "snv":
            if gene1 not in snv_gene:
                snv_gene[gene1] = []
            snv_gene[gene1].append(rule)
    output["snv_gene"] = snv_gene
    output["cnv_gene"] = cnv_gene
    output["fusion_gene"] = fusion_gene
    output["exp_gene"] = exp_gene
    return output


def process_files(patient_id, study_type):
    exp_gene = {}
    cnv_gene = {}
    fusion_gene = {}
    snv_gene = {}
    data_out_fusion = []
    data_out_cnv = []
    print(study_type)
    if study_type == "plan":
        workbook = load_workbook(ref_folder + "RULE BASE PLAN 2019.xlsx")
        all_rules = process_rules_plan(workbook)
        exp_gene = all_rules["exp_gene"]
        cnv_gene = all_rules["cnv_gene"]
        snv_gene = all_rules["snv_gene"]
        fusion_gene = all_rules["fusion_gene"]
    elif study_type == "mgt9":
        workbook = load_workbook(ref_folder + "MGT9_rules_final.xlsx")
        all_rules = process_rules_mgt9(workbook)
        exp_gene = all_rules["exp_gene"]
        cnv_gene = all_rules["cnv_gene"]
        snv_gene = all_rules["snv_gene"]
        fusion_gene = all_rules["fusion_gene"]
    elif study_type == "peach":
        workbook = load_workbook(ref_folder + "Peach_Rules_12232021_FINAL.xlsx") #new rules
        all_rules = process_rules_peach(workbook)
        exp_gene = all_rules["exp_gene"]
        cnv_gene = all_rules["cnv_gene"]
        snv_gene = all_rules["snv_gene"]
        fusion_gene = all_rules["fusion_gene"]


    patient = patient_id


    cna_suffix = ".final.cna.seg.vcf"
    score_suffix = ".Score"
    snv_suffix = ".snpeff.final.vcf"
    fusion_suffix = ".thf.vcf"


    print("exp")
    print(exp_gene)
    print("cnv")
    print(cnv_gene)
    print("fusion")
    print(fusion_gene)
    print("snv")
    print(snv_gene)

    data_out = get_score_genes(patient, exp_gene, score_suffix, 2, 0.75, 0.25)
    if study_type != "peach":
        data_out_snv = get_snp_genes(patient, snv_gene, snv_suffix)
    else:
        data_out_snv = get_snp_genes_peach(patient, snv_gene)
    if study_type != "peach":
        data_out_cnv = get_cnv_genes(patient, cnv_gene, cna_suffix, 1)
    if study_type != "peach":
        data_out_fusion = get_fusion_genes(patient, fusion_gene, fusion_suffix)

    print("=====output=====")
    write_html_file(patient)

    f = open(output_folder + patient + ".html", "a")
    if len(data_out_snv) > 0:
        f.write("<h3>Single Nucleotide Variants</h3>")
        f.write("<table width=\"800\" border=\"0\">\n")
        f.write("\t\t<tr>" +
                "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
                "Drug</th>" +
                "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
                "Gene</th>" +
                "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
                "AA Change</th></tr>")
    counter = 0
    repeat_drugs = {}
    for drug in data_out_snv:
        for rule in data_out_snv[drug]:
            gene = rule["gene"]
            test = gene + "_" + drug
            if test in repeat_drugs:
                continue
            repeat_drugs[test] = ""
            f.write("<tr>")
            f.write("<td width=\"200\">" + drug + "</td>" + "<td width=\"200\">" + rule["gene"] + "</td>" +
                    "<td width=\"200\">" + str(rule["aa_change"]) + "</td>")
            f.write("</tr>")
    if len(data_out_snv) > 0:
        f.write("</table>")
        f.write("<p style=\"page-break-after: always;\">&nbsp;</p>")

    if len(data_out_fusion) > 0:
        f.write("<h3>Fusions</h3>")
        f.write("<table width=\"800\" border=\"0\">\n")
        f.write(
            "\t\t<tr>" +
            "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
            "Drug</th>" +
            "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
            "Gene</th>"
            "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
            "Fusion</th><tr>")

    for drug in data_out_fusion:
        for rule in data_out_fusion[drug]:
            f.write("<td width=\"200\">" + drug + "</td>" + "<td width=\"200\">" + rule["gene"] + "</td>" +
                    "<td width=\"200\">" + str(rule["fusion"]) + "</td></tr>")
            if len(data_out_fusion) > 0:
                f.write("</table>")
            f.write("<p style=\"page-break-after: always;\">&nbsp;</p>")

    if len(data_out_cnv) > 0:
        f.write("<h3>Copy Number Variants</h3>")
        f.write("<table width=\"800\" border=\"0\">\n")
        f.write(
            "\t\t<tr>" +
            "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
            "Drug</th>" +
            "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
            "Gene</th>"
            "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
            "Log Fold Change</th></tr>")
        counter = 0
        for drug in data_out_cnv:
            for rule in data_out_cnv[drug]:
                f.write("<tr>")
                f.write("<td width=\"200\">" + drug + "</td>" + "<td width=\"200\">" + rule["gene"] + "</td>" +
                        "<td width=\"200\">" + str(rule["log_fold_change"]) + "</td></tr>")
    if len(data_out_cnv) > 0:
        f.write("</table>")
        f.write("<p style=\"page-break-after: always;\">&nbsp;</p>")

    print(data_out)
    f.write("<h3>Gene Expression Changes</h3>")
    drug_written = {}
    for drug in sorted(data_out):
        f.write("<table width=\"800\" border=\"0\">\n")
        for rule in data_out[drug]:
            if drug not in drug_written:
                f.write("<h3>" + drug + "</h3>")
                drug_written[drug] = ""
            f.write(
                "\t\t<tr>" +
                "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
                "Drug</th>" +
                "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
                "Gene</th>" +
                "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
                "Over/Under Expression</th>" +
                "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
                "Sensitive/Resistant</th>" +
                "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
                "NRZ</th>" +
                "<th style=\"font-family:Georgia, Garamond, Serif;color:white;\" border=\"0\" bgcolor=\"#4267B2\" align=left>" +
                "CRC</th></tr>")
            f.write("<tr>")
            f.write("<td width=\"200\">" + drug.replace("/","</br>")  + "</td>" + "<td width=\"200\">" + rule["gene"] + "</td>" +
                    "<td width=\"200\">" + str(rule["direction"]).upper() + "</td>" +
                    "<td width=\"200\">" + str(rule["sense"]) + "</td>" +
                    "<td width=\"200\">" + str(rule["z_score"]) + "</td>" +
                    "<td width=\"200\">" + str(rule["crc"]) + "</td></tr>")
        f.write("</table>")

    f.write("\t</body>\n")
    f.write("</html>\n")
    f.flush()
    f.close()
    # print(data_out_fusion)
    # print(data_out)
    # print(data_out_snv)
    # print(data_out_cnv)


@app.route("/")
def home():
    html_file = open("./static/index.html","r")
    output = html_file.read()
    return output


@app.route("/new_patient")
def upload_file():
    webpage = "<html><form method=\"POST\" enctype=\"multipart/form-data\" action=\"upload\">"
    webpage = webpage + "<label for=\"patient\">Patient ID:</label>"
    webpage = webpage + "<input type=\"text\" id=\"patient\" name=\"patient\"><br><br>"
    webpage = webpage + "<input type=\"file\" name=\"file\" multiple=\"\"><br><br>"
    webpage = webpage + "<select name=\"study_type\" id=\"study_type\">"
    webpage = webpage + "<option value=\"plan\">plan</option>"
    webpage = webpage + "<option value=\"mgt9\">mgt9</option>"
    webpage = webpage + "<option value=\"peach\">peach</option>"
    webpage = webpage + "</select>"
    webpage = webpage + "<input type=\"submit\" value=\"add\">"
    webpage = webpage + "</form>"
    return webpage

@app.route("/existing_report")
def get_report():
    webpage = "<html>"
    webpage = webpage + "<form method=\"POST\" action=\"/process_existing\">"
    webpage = webpage + "<select name=\"patient\" size=\"10\">"
    for file in os.listdir("./homedir/"):
        if file.endswith(".html"):
            print(file)
            webpage = webpage + "<option value=\"" + file + "\">" + file.replace(".html", "") + "</option>"
    webpage = webpage + "</select>"
    webpage = webpage + "<input type=\"submit\" value=\"Get Report\">"
    webpage = webpage + "</form>"
    webpage = webpage + "</html>"
    #print(webpage)
    return webpage


@app.route('/process_existing', methods=['GET','POST'])
def get_existing():
    file = request.form.get("patient")
    print(file)
    output_file = "./homedir/" + file
    read_file = open(output_file,"r")
    webpage = read_file.read()
    return webpage

@app.route('/upload', methods=['GET','POST'])
def upload():
    print("Here 1")
    file_names = ""
    if not _upload_dir:
        raise ValueError('Uploads are disabled.')
    patient = request.form['patient']
    report_type = request.form['study_type']
    print(report_type)
    if request.method == "POST":
        files = request.files.getlist("file")
        print("Here")
        for file in files:
            print(file.filename)
            file.save(os.path.join(_upload_dir, file.filename))
    process_files(patient, report_type)
    f = open(output_folder + patient + ".html", "r")
    blah = f.read()
    print(app.root_path)
    return blah


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0",port=int(os.getenv('PORT', 80)))



